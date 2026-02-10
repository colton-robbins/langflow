"""Stop-Loss template filler component.

Copies an Excel template and fills the Stop-Loss sheet with data from carrier JSON files.
Supports QBE (renewal_options) and Berkley (specific_stop_loss_options) JSON formats.
"""

from __future__ import annotations

import base64
import io
import json
import os
import re
import shutil
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from lfx.base.data.storage_utils import read_file_bytes, read_file_text
from lfx.custom.custom_component.component import Component
from lfx.io import BoolInput, FileInput, HandleInput, IntInput, Output
from lfx.schema.data import Data
from lfx.schema.message import Message
from lfx.utils.async_helpers import run_until_complete

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class DataQualityReport:
    """Track missing and problematic data during processing."""

    def __init__(self):
        self.issues: dict[str, list[str]] = defaultdict(list)
        self.carrier_context: str | None = None
        self.option_context: str | int | None = None

    def set_carrier(self, carrier_name: str) -> None:
        self.carrier_context = carrier_name
        self.option_context = None

    def set_option(self, option_number: str | int) -> None:
        self.option_context = option_number

    def add_missing(self, field_name: str, fallback_value: object = None) -> None:
        context = self._get_context()
        msg = f"Missing: {field_name}"
        if fallback_value is not None:
            msg += f" (using default: {fallback_value})"
        self.issues[context].append(msg)

    def add_warning(self, message: str) -> None:
        context = self._get_context()
        self.issues[context].append(f"Warning: {message}")

    def add_error(self, message: str) -> None:
        context = self._get_context()
        self.issues[context].append(f"ERROR: {message}")

    def _get_context(self) -> str:
        if self.option_context is not None:
            return f"{self.carrier_context} - Option {self.option_context}"
        if self.carrier_context:
            return f"{self.carrier_context} - General"
        return "General"

    def has_issues(self) -> bool:
        return len(self.issues) > 0

    def to_string(self) -> str:
        if not self.has_issues():
            return "DATA QUALITY REPORT: No issues found - all data complete!"
        lines = ["DATA QUALITY REPORT: Missing or Problematic Data"]
        for context, issue_list in sorted(self.issues.items()):
            lines.append(f"\n{context}:")
            for issue in issue_list:
                lines.append(f"  - {issue}")
        lines.append(f"\nTotal contexts with issues: {len(self.issues)}")
        lines.append(f"Total issues found: {sum(len(v) for v in self.issues.values())}")
        return "\n".join(lines)


def _normalize_carrier_data(
    data: dict,
    carrier_name: str,
    report: DataQualityReport | None = None,
) -> dict:
    """Normalize different carrier JSON structures to a common format."""
    normalized: dict = {
        "carrier_name": carrier_name,
        "options": [],
    }

    if "options" in data and "enrollment" in data:
        normalized["carrier_name"] = data.get("carrier_name", carrier_name)
        normalized["tpa"] = data.get("tpa", "MERITAIN")
        if "tpa" not in data and report:
            report.add_missing("tpa", "MERITAIN")
        normalized["aggregate_corridor"] = data.get("aggregate_corridor", "125%")
        if "aggregate_corridor" not in data and report:
            report.add_missing("aggregate_corridor", "125%")
        normalized["enrollment"] = data.get("enrollment", {})
        if "enrollment" not in data and report:
            report.add_missing("enrollment")

        for option in data["options"]:
            option_num = option.get("option_number", "Unknown")
            if report:
                report.set_option(option_num)

            specific = option.get("specific_stop_loss", {})
            aggregate = option.get("aggregate_stop_loss", {})

            if not specific:
                if report:
                    report.add_error("Missing specific_stop_loss section")
                continue
            if not aggregate:
                if report:
                    report.add_error("Missing aggregate_stop_loss section")
                continue

            if "specific_deductible" not in specific and report:
                report.add_missing("specific_deductible", 0)
            if "contract_basis" not in specific and report:
                report.add_missing("contract_basis (specific)", "Unknown")
            if "contract_basis" not in aggregate and report:
                report.add_missing("contract_basis (aggregate)", "Unknown")
            if "coverages" not in specific and report:
                report.add_missing("coverages", "Unknown")
            if "composite_rate_per_month" not in aggregate and report:
                report.add_missing("composite_rate_per_month", 0)
            if "aggregate_deductible" not in aggregate and report:
                report.add_missing("aggregate_deductible", 0)

            norm_option: dict = {
                "option_number": option.get("option_number", "Unknown"),
                "specific_deductible": specific.get("specific_deductible", 0),
                "contract_type_specific": specific.get("contract_basis", "Unknown"),
                "contract_type_aggregate": aggregate.get("contract_basis", "Unknown"),
                "coverages": specific.get("coverages", "Unknown"),
                "rates": {},
                "aggregate_rate": aggregate.get("composite_rate_per_month", 0),
                "aggregate_factors": {},
                "aggregate_deductible": aggregate.get("aggregate_deductible", 0),
                "individual_exceptions": specific.get("individual_exceptions", []),
                "run_in_limit_aggregate": aggregate.get("run_in_limit"),
            }

            rates_extracted = False
            if "monthly_rates" in specific:
                monthly_rates = specific["monthly_rates"]
                for tier_key, tier_name in [
                    ("employee", "Employee"),
                    ("employee_plus_spouse", "Employee plus Spouse"),
                    ("employee_plus_children", "Employee plus Child(ren)"),
                    ("family", "Family"),
                ]:
                    if tier_key not in monthly_rates and report:
                        report.add_missing(f"monthly_rates.{tier_key}", 0)
                    norm_option["rates"][tier_name] = monthly_rates.get(tier_key, 0)
                rates_extracted = True
            elif "monthly_premium_and_fees" in specific:
                monthly_premium = specific["monthly_premium_and_fees"]
                for tier_key, tier_name in [
                    ("employee", "Employee"),
                    ("employee_plus_spouse", "Employee plus Spouse"),
                    ("employee_plus_children", "Employee plus Child(ren)"),
                    ("family", "Family"),
                ]:
                    if tier_key not in monthly_premium and report:
                        report.add_missing(f"monthly_premium_and_fees.{tier_key}", 0)
                    tier_data = monthly_premium.get(tier_key, {})
                    if isinstance(tier_data, dict) and "premium" not in tier_data and report:
                        report.add_missing(f"monthly_premium_and_fees.{tier_key}.premium", 0)
                    norm_option["rates"][tier_name] = monthly_premium.get(tier_key, {}).get("premium", 0)
                rates_extracted = True

            if not rates_extracted and report:
                report.add_error("No recognized rate structure (expected monthly_rates or monthly_premium_and_fees)")

            factors_extracted = False
            if "monthly_aggregate_claim_factors" in aggregate:
                agg_factors = aggregate["monthly_aggregate_claim_factors"]
                if isinstance(agg_factors.get("employee"), (int, float)):
                    for tier_key, tier_name in [
                        ("employee", "Employee"),
                        ("employee_plus_spouse", "Employee plus Spouse"),
                        ("employee_plus_children", "Employee plus Child(ren)"),
                        ("family", "Family"),
                    ]:
                        if tier_key not in agg_factors and report:
                            report.add_missing(f"monthly_aggregate_claim_factors.{tier_key}", 0)
                        norm_option["aggregate_factors"][tier_name] = agg_factors.get(tier_key, 0)
                    factors_extracted = True
                elif isinstance(agg_factors.get("employee"), dict):
                    for tier_key, tier_name in [
                        ("employee", "Employee"),
                        ("employee_plus_spouse", "Employee plus Spouse"),
                        ("employee_plus_children", "Employee plus Child(ren)"),
                        ("family", "Family"),
                    ]:
                        if tier_key not in agg_factors and report:
                            report.add_missing(f"monthly_aggregate_claim_factors.{tier_key}", 0)
                        tier_data = agg_factors.get(tier_key, {})
                        if isinstance(tier_data, dict) and "factor" not in tier_data and report:
                            report.add_missing(f"monthly_aggregate_claim_factors.{tier_key}.factor", 0)
                        norm_option["aggregate_factors"][tier_name] = agg_factors.get(tier_key, {}).get("factor", 0)
                    factors_extracted = True

            if not factors_extracted and report:
                report.add_error("No recognized aggregate factor structure or missing monthly_aggregate_claim_factors")

            normalized["options"].append(norm_option)

        return normalized

    if "renewal_options" in data:
        normalized["carrier_name"] = data.get("proposal_information", {}).get("carrier", "QBE")
        normalized["tpa"] = data.get("plan_assumptions", {}).get("ur_vendor", "MERITAIN")
        normalized["aggregate_corridor"] = data.get("plan_assumptions", {}).get("aggregate_risk_corridor", "125%")
        normalized["enrollment"] = data.get("total_enrollment_summary", {}).get("aggregate_coverage", {})

        for option in data.get("renewal_options", []):
            specific = option.get("specific_stop_loss", {})
            aggregate = option.get("aggregate_stop_loss", {})
            if not specific or not aggregate:
                continue

            agg_rate = 0
            if "rate_per_month" in aggregate:
                rate_info = aggregate["rate_per_month"]
                agg_rate = rate_info.get("composite_rate", 0) if isinstance(rate_info, dict) else rate_info

            norm_option = {
                "option_number": option.get("option_number", "Unknown"),
                "specific_deductible": specific.get("specific_deductible_per_individual", specific.get("specific_deductible", 0)),
                "contract_type_specific": specific.get("contract_basis", "Unknown"),
                "contract_type_aggregate": aggregate.get("contract_basis", "Unknown"),
                "coverages": specific.get("coverages", "Unknown"),
                "rates": {},
                "aggregate_rate": agg_rate,
                "aggregate_factors": {},
                "aggregate_deductible": aggregate.get("aggregate_deductible", 0),
                "individual_exceptions": specific.get("individual_exceptions", []),
                "run_in_limit_aggregate": aggregate.get("run_in_limit"),
            }

            if "rates_per_month" in specific:
                for rate_info in specific["rates_per_month"]:
                    tier = rate_info.get("tier", "")
                    if tier and tier != "Composite":
                        norm_option["rates"][tier] = rate_info.get("rate", 0)

            if "monthly_aggregate_claim_factors" in aggregate:
                for factor_info in aggregate["monthly_aggregate_claim_factors"]:
                    tier = factor_info.get("tier", "")
                    if tier and tier != "Composite":
                        norm_option["aggregate_factors"][tier] = factor_info.get("factor", 0)

            normalized["options"].append(norm_option)

    elif "specific_stop_loss_options" in data:
        normalized["carrier_name"] = data.get("insurance_carrier", {}).get("name", "BERKLEY")
        normalized["tpa"] = data.get("plan_administration", {}).get("claims_administrator", "MERITAIN")
        agg_options = data.get("aggregate_stop_loss_options", [])
        normalized["aggregate_corridor"] = agg_options[0].get("aggregate_corridor", "125%") if agg_options else "125%"
        normalized["enrollment"] = data.get("enrollment_summary", {}).get("aggregate_breakdown", {})

        specific_options = data.get("specific_stop_loss_options", [])
        aggregate_options = data.get("aggregate_stop_loss_options", [])

        for idx, specific_opt in enumerate(specific_options):
            if idx >= len(aggregate_options):
                continue
            aggregate_opt = aggregate_options[idx]

            agg_rate = 0
            if "composite_rate_per_month" in aggregate_opt:
                rate_info = aggregate_opt["composite_rate_per_month"]
                agg_rate = rate_info.get("rate", 0) if isinstance(rate_info, dict) else rate_info

            norm_option = {
                "option_number": specific_opt.get("option", idx + 1),
                "specific_deductible": specific_opt.get("annual_specific_deductible_per_individual", specific_opt.get("specific_deductible", 0)),
                "contract_type_specific": specific_opt.get("contract_type", "Unknown"),
                "contract_type_aggregate": aggregate_opt.get("contract_type", "Unknown"),
                "coverages": specific_opt.get("coverages", "Unknown"),
                "rates": {},
                "aggregate_rate": agg_rate,
                "aggregate_factors": {},
                "aggregate_deductible": aggregate_opt.get("annual_aggregate_deductible", aggregate_opt.get("aggregate_deductible", 0)),
                "individual_exceptions": specific_opt.get("individual_exceptions", []),
                "run_in_limit_aggregate": aggregate_opt.get("run_in_limited_to", aggregate_opt.get("run_in_limit")),
            }

            if "monthly_rates" in specific_opt:
                monthly_rates = specific_opt["monthly_rates"]
                if isinstance(monthly_rates.get("employee"), dict):
                    norm_option["rates"]["Employee"] = monthly_rates.get("employee", {}).get("rate", 0)
                    norm_option["rates"]["Employee plus Spouse"] = monthly_rates.get("employee_plus_spouse", {}).get("rate", 0)
                    norm_option["rates"]["Employee plus Child(ren)"] = monthly_rates.get("employee_plus_children", {}).get("rate", 0)
                    norm_option["rates"]["Family"] = monthly_rates.get("family", {}).get("rate", 0)
                else:
                    norm_option["rates"]["Employee"] = monthly_rates.get("employee", 0)
                    norm_option["rates"]["Employee plus Spouse"] = monthly_rates.get("employee_plus_spouse", 0)
                    norm_option["rates"]["Employee plus Child(ren)"] = monthly_rates.get("employee_plus_children", 0)
                    norm_option["rates"]["Family"] = monthly_rates.get("family", 0)

            if "monthly_aggregate_claim_factors" in aggregate_opt:
                agg_factors = aggregate_opt["monthly_aggregate_claim_factors"]
                if isinstance(agg_factors.get("employee"), dict):
                    norm_option["aggregate_factors"]["Employee"] = agg_factors.get("employee", {}).get("factor", 0)
                    norm_option["aggregate_factors"]["Employee plus Spouse"] = agg_factors.get("employee_plus_spouse", {}).get("factor", 0)
                    norm_option["aggregate_factors"]["Employee plus Child(ren)"] = agg_factors.get("employee_plus_children", {}).get("factor", 0)
                    norm_option["aggregate_factors"]["Family"] = agg_factors.get("family", {}).get("factor", 0)
                else:
                    norm_option["aggregate_factors"]["Employee"] = agg_factors.get("employee", 0)
                    norm_option["aggregate_factors"]["Employee plus Spouse"] = agg_factors.get("employee_plus_spouse", 0)
                    norm_option["aggregate_factors"]["Employee plus Child(ren)"] = agg_factors.get("employee_plus_children", 0)
                    norm_option["aggregate_factors"]["Family"] = agg_factors.get("family", 0)

            normalized["options"].append(norm_option)

    return normalized


def _fill_option_data(
    ws: object,
    option_data: dict,
    col: int,
    carrier_info: dict,
    report: DataQualityReport | None = None,
) -> None:
    """Fill a single option's data into a specific column."""
    tier_row_map = {"Employee": 41, "Employee plus Spouse": 44, "Employee plus Child(ren)": 47, "Family": 50}
    factor_row_map = {"Employee": 84, "Employee plus Spouse": 85, "Employee plus Child(ren)": 86, "Family": 87}
    expected_claims_map = {"Employee": 74, "Employee plus Spouse": 75, "Employee plus Child(ren)": 76, "Family": 77}

    ws.cell(row=20, column=col).value = carrier_info.get("carrier_name", "UNKNOWN").split()[0].upper()
    ws.cell(row=21, column=col).value = "DIRECT"
    ws.cell(row=22, column=col).value = carrier_info.get("tpa", "MERITAIN").upper()
    ws.cell(row=24, column=col).value = option_data.get("specific_deductible", 0)
    ws.cell(row=25, column=col).value = 0
    ws.cell(row=26, column=col).value = option_data.get("contract_type_specific", "Unknown")
    ws.cell(row=27, column=col).value = option_data.get("contract_type_aggregate", "Unknown")
    coverages = str(option_data.get("coverages", "Unknown")).replace("Rx Card", "RX")
    ws.cell(row=28, column=col).value = coverages
    ws.cell(row=29, column=col).value = coverages
    ws.cell(row=30, column=col).value = 0
    run_in_limit = option_data.get("run_in_limit_aggregate")
    if run_in_limit is not None:
        ws.cell(row=31, column=col).value = run_in_limit
    ws.cell(row=32, column=col).value = 0
    ws.cell(row=33, column=col).value = "N"
    ws.cell(row=34, column=col).value = "Y"
    ws.cell(row=35, column=col).value = 0

    rates = option_data.get("rates", {})
    for tier, rate in rates.items():
        if tier in tier_row_map:
            ws.cell(row=tier_row_map[tier], column=col).value = rate
    if not rates and report:
        report.add_warning("No rates found for this option")

    ws.cell(row=59, column=col).value = option_data.get("aggregate_rate", 0)

    aggregate_factors = option_data.get("aggregate_factors", {})
    for tier, factor in aggregate_factors.items():
        if tier in factor_row_map:
            ws.cell(row=factor_row_map[tier], column=col).value = factor
    if not aggregate_factors and report:
        report.add_warning("No aggregate factors found for this option")

    corridor_str = carrier_info.get("aggregate_corridor", "125%")
    try:
        corridor_value = float(str(corridor_str).strip("%")) / 100
    except (ValueError, AttributeError):
        corridor_value = 1.25
        if report:
            report.add_warning(f"Invalid corridor value '{corridor_str}', using default 1.25")

    ws.cell(row=94, column=col).value = corridor_value

    if aggregate_factors and corridor_value > 0:
        for tier, factor in aggregate_factors.items():
            if tier in expected_claims_map:
                try:
                    expected_value = factor / corridor_value
                    ws.cell(row=expected_claims_map[tier], column=col).value = expected_value
                except (TypeError, ZeroDivisionError) as e:
                    if report:
                        report.add_warning(f"Could not calculate expected cost for {tier}: {e}")

    ws.cell(row=95, column=col).value = option_data.get("aggregate_deductible", 0)

    individual_exceptions = option_data.get("individual_exceptions", [])
    standard_deductible = option_data.get("specific_deductible", 0)
    for idx, exception in enumerate(individual_exceptions[:2]):
        try:
            exception_deductible = exception.get("deductible", 0)
            laser_diff = exception_deductible - standard_deductible
            ws.cell(row=98 + idx, column=col).value = laser_diff
        except Exception as e:
            if report:
                report.add_warning(f"Could not process individual exception #{idx + 1}: {e}")


def _recalculate_formulas(filepath: str) -> None:
    """Evaluate formulas and inject cached values. Requires formulas and numpy."""
    try:
        import formulas
        import numpy as np
    except ImportError:
        return

    try:
        import schedula as sh
        import zipfile
        import xml.etree.ElementTree as ET
        import io as io_module
    except ImportError:
        return

    from formulas.tokens.operand import XlError

    SSML = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"

    tmp_dir = tempfile.mkdtemp()
    tmp_path = os.path.join(tmp_dir, "calc_temp.xlsx")
    shutil.copy(filepath, tmp_path)

    unsupported_patterns = ["[#This Row]", "XLOOKUP", "_xlfn.", "_xludf."]
    tmp_wb = openpyxl.load_workbook(tmp_path)
    for ws in tmp_wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    if any(p in cell.value for p in unsupported_patterns):
                        cell.value = None
    tmp_wb.save(tmp_path)

    xl_model = formulas.ExcelModel().loads(tmp_path).finish()
    solution = xl_model.calculate()

    os.remove(tmp_path)
    os.rmdir(tmp_dir)

    cache: dict[str, dict[str, object]] = {}
    for key, result in solution.items():
        if not hasattr(result, "ranges") or not result.ranges:
            continue
        rng_meta = result.ranges[0]
        sheet_upper = rng_meta.get("sheet", "").upper()
        if not sheet_upper:
            continue
        if sheet_upper not in cache:
            cache[sheet_upper] = {}
        ref = rng_meta.get("ref", "")
        flat_values = np.ravel(result.value)
        if ":" in ref:
            match = re.match(r"([A-Z]+)(\d+):([A-Z]+)(\d+)", ref, re.IGNORECASE)
            if match:
                from openpyxl.utils import column_index_from_string
                c1 = column_index_from_string(match.group(1))
                r1 = int(match.group(2))
                c2 = column_index_from_string(match.group(3))
                r2 = int(match.group(4))
                idx = 0
                for r in range(r1, r2 + 1):
                    for c in range(c1, c2 + 1):
                        if idx < len(flat_values):
                            cell_key = f"{get_column_letter(c)}{r}"
                            cache[sheet_upper][cell_key] = flat_values[idx]
                            idx += 1
        else:
            if len(flat_values) == 1:
                cache[sheet_upper][ref.upper()] = flat_values[0]

    def to_xml_value(val: object) -> tuple[str, str] | None:
        if hasattr(val, "value") and hasattr(val.value, "shape"):
            if val.value.shape == (1, 1):
                val = val.value[0, 0]
        if val is getattr(sh, "EMPTY", None) or val is None:
            return None
        if isinstance(val, XlError):
            return ("e", str(val))
        if isinstance(val, np.generic):
            val = val.item()
        if isinstance(val, bool):
            return ("b", "1" if val else "0")
        if isinstance(val, (int, float)):
            if isinstance(val, float) and (np.isnan(val) or np.isinf(val)):
                return None
            return ("n", repr(val))
        if isinstance(val, str):
            return ("str", val)
        return None

    ns_r = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    with zipfile.ZipFile(filepath, "r") as zin:
        wb_xml = ET.parse(zin.open("xl/workbook.xml")).getroot()
        rels_xml = ET.parse(zin.open("xl/_rels/workbook.xml.rels")).getroot()
        rid_to_target = {rel.get("Id"): rel.get("Target") for rel in rels_xml}
        sheet_to_xml = {}
        for sheet_el in wb_xml.iter(f"{{{SSML}}}sheet"):
            name = sheet_el.get("name", "")
            rid = sheet_el.get(f"{{{ns_r}}}id", "")
            target = rid_to_target.get(rid, "")
            if target:
                xml_path = target.lstrip("/")
                if not xml_path.startswith("xl/"):
                    xml_path = "xl/" + xml_path
                sheet_to_xml[name.upper()] = xml_path

        file_contents = {item.filename: zin.read(item.filename) for item in zin.infolist()}

    ET.register_namespace("", SSML)
    for sheet_upper, values in cache.items():
        xml_path = sheet_to_xml.get(sheet_upper)
        if not xml_path or xml_path not in file_contents:
            continue
        raw_xml = file_contents[xml_path].decode("utf-8")
        for ns_match in re.finditer(r'xmlns:(\w+)="([^"]+)"', raw_xml[:2000]):
            ET.register_namespace(ns_match.group(1), ns_match.group(2))
        tree = ET.ElementTree(ET.fromstring(file_contents[xml_path]))
        root = tree.getroot()
        for c_elem in root.iter(f"{{{SSML}}}c"):
            f_elem = c_elem.find(f"{{{SSML}}}f")
            if f_elem is None:
                continue
            ref = c_elem.get("r", "").upper()
            if ref not in values:
                continue
            result = to_xml_value(values[ref])
            if result is None:
                continue
            type_attr, text_val = result
            v_elem = c_elem.find(f"{{{SSML}}}v")
            if v_elem is None:
                v_elem = ET.SubElement(c_elem, f"{{{SSML}}}v")
            v_elem.text = text_val
            if type_attr == "str":
                c_elem.set("t", "str")
            elif type_attr == "b":
                c_elem.set("t", "b")
            elif type_attr == "e":
                c_elem.set("t", "e")
            elif "t" in c_elem.attrib:
                del c_elem.attrib["t"]
        buf = io.BytesIO()
        tree.write(buf, xml_declaration=True, encoding="UTF-8")
        file_contents[xml_path] = buf.getvalue()

    with zipfile.ZipFile(filepath, "w", zipfile.ZIP_DEFLATED) as zout:
        for filename, data in file_contents.items():
            zout.writestr(filename, data)


class StopLossTemplateFillerComponent(Component):
    """Fill Stop-Loss Excel template with carrier JSON data."""

    display_name = "Stop-Loss Template Filler"
    description = (
        "Fills an Excel Stop-Loss template with data from one or more carrier JSON files. "
        "Supports QBE (renewal_options) and Berkley (specific_stop_loss_options) formats."
    )
    icon = "file-spreadsheet"
    name = "StopLossTemplateFiller"

    inputs = [
        FileInput(
            name="excel_template",
            display_name="Excel Template",
            file_types=["xlsx"],
            info="Clean Excel template with Stop-Loss sheet to fill.",
            required=False,
        ),
        HandleInput(
            name="excel_template_input",
            display_name="Excel Template (from input)",
            input_types=["Data", "Message"],
            info="Excel template from upstream component (e.g. Read File). Data must have file_path.",
            required=False,
        ),
        FileInput(
            name="json_files",
            display_name="Carrier JSON Files",
            file_types=["json"],
            info="One or more carrier JSON data files (QBE or Berkley format).",
            required=False,
            is_list=True,
        ),
        HandleInput(
            name="json_files_input",
            display_name="Carrier JSON Files (from input)",
            input_types=["Data", "Message"],
            info="Carrier JSON files from upstream component. Data/Message must have file_path.",
            required=False,
            is_list=True,
        ),
        IntInput(
            name="first_carrier_column",
            display_name="First Carrier Column",
            value=8,
            info="Starting column for first carrier (8 = H). Subsequent carriers use +4 columns.",
            advanced=True,
        ),
        BoolInput(
            name="recalculate_formulas",
            display_name="Recalculate Formulas",
            value=True,
            info="Evaluate workbook formulas and inject cached values (requires formulas library).",
            advanced=True,
        ),
    ]

    outputs = [
        Output(
            display_name="Filled Workbook",
            name="workbook_data",
            method="fill_template",
            types=["Data"],
        ),
        Output(
            display_name="Quality Report",
            name="quality_report",
            method="get_quality_report",
            types=["Message"],
        ),
    ]

    def __init__(self, **kwargs) -> None:
        super().__init__(**kwargs)
        self._report = DataQualityReport()

    def _extract_path_from_input(self, obj: Data | Message) -> str | None:
        """Extract file path from Data or Message."""
        if isinstance(obj, Data):
            return obj.data.get("file_path") if isinstance(obj.data, dict) else None
        if isinstance(obj, Message):
            text = obj.text
            return str(text).strip() if text else None
        return None

    def _get_template_path(self) -> str | None:
        """Get template path from FileInput or HandleInput."""
        input_val = getattr(self, "excel_template_input", None)
        if input_val:
            if isinstance(input_val, list):
                input_val = input_val[0] if input_val else None
            if input_val:
                path = self._extract_path_from_input(input_val)
                if path:
                    return path
        if self.excel_template:
            return self.excel_template if isinstance(self.excel_template, str) else None
        return None

    def _get_json_paths(self) -> list[str]:
        """Get JSON file paths from FileInput or HandleInput."""
        paths: list[str] = []
        input_val = getattr(self, "json_files_input", None)
        if input_val:
            for obj in (input_val if isinstance(input_val, list) else [input_val]):
                if obj:
                    p = self._extract_path_from_input(obj)
                    if p:
                        paths.append(p)
        if paths:
            return paths
        json_files = getattr(self, "json_files", None)
        if json_files:
            paths.extend([p for p in (json_files if isinstance(json_files, list) else [json_files]) if p])
        return paths

    def _resolve_and_read_text(self, path: str) -> str:
        """Read file text from storage or local path."""
        return run_until_complete(
            read_file_text(path, encoding="utf-8", resolve_path=self.resolve_path)
        )

    def _load_template_bytes(self, path: str) -> bytes:
        """Load template file bytes from storage or local path."""
        return run_until_complete(read_file_bytes(path, resolve_path=self.resolve_path))

    def fill_template(self) -> Data:
        """Fill the template with JSON data and return the workbook as Data."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required. Install with: pip install openpyxl")

        template_path = self._get_template_path()
        if not template_path:
            raise ValueError(
                "Excel template is required. Provide via 'Excel Template' upload or "
                "'Excel Template (from input)' from an upstream component."
            )

        json_paths = self._get_json_paths()
        if not json_paths:
            raise ValueError(
                "At least one JSON file is required. Provide via 'Carrier JSON Files' upload "
                "or 'Carrier JSON Files (from input)' from an upstream component."
            )

        self._report = DataQualityReport()
        self._fill_has_run = False

        template_bytes = self._load_template_bytes(template_path)

        output_dir = tempfile.mkdtemp()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = os.path.join(output_dir, f"filled_template_{timestamp}.xlsx")

        try:
            with open(output_file, "wb") as f:
                f.write(template_bytes)

            wb = openpyxl.load_workbook(output_file)
            if "Stop-Loss" not in wb.sheetnames:
                raise ValueError("Template must contain a 'Stop-Loss' sheet")
            ws = wb["Stop-Loss"]

            first_col = getattr(self, "first_carrier_column", 8) or 8
            enrollment_filled = False

            for idx, json_path in enumerate(json_paths):
                carrier_name = f"Carrier {idx + 1}"
                self._report.set_carrier(carrier_name)

                try:
                    json_text = self._resolve_and_read_text(json_path)
                except Exception as e:
                    self._report.add_error(f"Failed to read {json_path}: {e}")
                    continue

                try:
                    raw_data = json.loads(json_text)
                except json.JSONDecodeError as e:
                    self._report.add_error(f"Invalid JSON in {json_path}: {e}")
                    continue

                try:
                    carrier_data = _normalize_carrier_data(raw_data, carrier_name, self._report)
                except Exception as e:
                    self._report.add_error(f"Failed to normalize carrier data: {e}")
                    continue

                if not carrier_data.get("options"):
                    self._report.add_warning(f"No options found for {carrier_name}, skipping")
                    continue

                if not enrollment_filled:
                    enrollment = carrier_data.get("enrollment", {})
                    if enrollment:
                        ws.cell(row=7, column=4).value = enrollment.get("employee_only", enrollment.get("employee", 0))
                        ws.cell(row=8, column=4).value = enrollment.get("employee_plus_spouse", 0)
                        ws.cell(row=9, column=4).value = enrollment.get("employee_plus_children", 0)
                        ws.cell(row=10, column=4).value = enrollment.get("family", 0)
                        enrollment_filled = True
                    else:
                        self._report.add_warning("No enrollment data found")

                start_column = first_col + idx * 4
                for opt_idx, option in enumerate(carrier_data["options"]):
                    col = start_column + opt_idx
                    self._report.set_option(option.get("option_number", opt_idx + 1))
                    try:
                        _fill_option_data(ws, option, col, carrier_data, self._report)
                    except Exception as e:
                        self._report.add_error(f"Failed to fill option: {e}")

            wb.calculation.calcMode = "auto"
            wb.calculation.fullCalcOnLoad = True
            wb.save(output_file)

            if getattr(self, "recalculate_formulas", True):
                try:
                    _recalculate_formulas(output_file)
                except Exception:
                    pass

            with open(output_file, "rb") as f:
                workbook_bytes = f.read()

            filename = f"filled_template_{timestamp}.xlsx"
            self.status = f"Template filled successfully. Output: {filename}"
            self._fill_has_run = True
            return Data(
                data={
                    "content_base64": base64.b64encode(workbook_bytes).decode("ascii"),
                    "filename": filename,
                    "mime_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                }
            )
        finally:
            try:
                shutil.rmtree(output_dir, ignore_errors=True)
            except OSError:
                pass

    def get_quality_report(self) -> Message:
        """Return the data quality report from the last fill operation."""
        if not getattr(self, "_fill_has_run", False):
            try:
                self.fill_template()
            except Exception as e:
                return Message(text=f"Fill failed before report: {e}")
        report = getattr(self, "_report", None)
        if report is None:
            return Message(text="No report available. Run fill_template first.")
        return Message(text=report.to_string())
